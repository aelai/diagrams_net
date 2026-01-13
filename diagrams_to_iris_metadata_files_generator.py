#!/usr/bin/env python3
"""
diagrams_to_iris_metadata_files_generator.py

V1: Generate IRiS Source + Target + Mapping metadata XLSX from a diagrams.net (draw.io) XML export.

Folder convention (relative to this script):
  ./model   -> input draw.io XML files (one or many)
  ./output  -> generated artifacts
    ./output/<MODEL_NAME>/source_<MODEL_NAME>.xlsx
    ./output/<MODEL_NAME>/target_<MODEL_NAME>.xlsx
    ./output/<MODEL_NAME>/mapping_<MODEL_NAME>.xlsx

MODEL_NAME is the input filename stem (filename without extensions).

Usage examples:
  python diagrams_to_iris_metadata_files_generator.py --schema landing --model SERVICE_SERVICESSYSTEM.drawio.xml
  python diagrams_to_iris_metadata_files_generator.py --schema landing --all
"""

from __future__ import annotations

import argparse
import base64
import dataclasses
import html
import json
import logging
import re
import urllib.parse
import zlib
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook


# ----------------------------- Data structures -----------------------------

@dataclasses.dataclass(frozen=True)
class SourceColumn:
    column: str
    datatype: str
    size: str
    scale: str
    # Optional: draw.io source-table “Target” column (e.g., h_service / s_xxx),
    # used later to help populate Target/Mapping. Not written to the Source XLSX.
    target: str = ""


@dataclasses.dataclass(frozen=True)
class SourceTable:
    table_name: str
    columns: List[SourceColumn]


# ----------------------------- Draw.io decoding ----------------------------

def _read_text(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="replace")


def _try_decode_diagram_payload(payload: str) -> Optional[str]:
    """
    Attempt to decode diagrams.net compressed diagram content.

    Common patterns:
      - Plain XML: '<mxGraphModel>...'
      - Base64 + raw DEFLATE (zlib -15) -> URL-encoded XML

    Returns decoded XML text if successful, else None.
    """
    s = (payload or "").strip()
    if not s:
        return None

    # Plain mxGraphModel or related XML
    if s.startswith("<mxGraphModel") or s.startswith("<mxfile") or s.startswith("<mxGraph"):
        return s

    # Sometimes mxGraphModel is embedded inline with other text
    if "<mxGraphModel" in s and "</mxGraphModel>" in s:
        start = s.find("<mxGraphModel")
        end = s.rfind("</mxGraphModel>") + len("</mxGraphModel>")
        return s[start:end]

    # Otherwise: try base64 decode + inflate
    try:
        raw = base64.b64decode(s, validate=False)
    except Exception:
        return None

    # Try raw DEFLATE first (most common), then zlib wrapper
    for wbits in (-15, zlib.MAX_WBITS):
        try:
            inflated = zlib.decompress(raw, wbits).decode("utf-8", errors="replace")
            inflated = urllib.parse.unquote(inflated)
            if "<mxGraphModel" in inflated and "</mxGraphModel>" in inflated:
                start = inflated.find("<mxGraphModel")
                end = inflated.rfind("</mxGraphModel>") + len("</mxGraphModel>")
                return inflated[start:end]
        except Exception:
            continue

    return None


def extract_mxgraphmodel_xml(drawio_xml_text: str) -> str:
    """
    Given a draw.io XML export (mxfile/diagram), return the embedded
    <mxGraphModel>...</mxGraphModel> XML text.

    Supports both compressed and uncompressed diagram payloads.

    Raises ValueError if not found/decodable.
    """
    import xml.etree.ElementTree as ET

    try:
        root = ET.fromstring(drawio_xml_text)
    except Exception as e:
        raise ValueError(f"Input is not valid XML: {e}") from e

    # Case 1: file already is mxGraphModel
    if root.tag == "mxGraphModel":
        return drawio_xml_text

    # Case 2: standard <mxfile><diagram>...</diagram></mxfile>
    if root.tag == "mxfile":
        diagram_elems = list(root.findall(".//diagram"))
        if not diagram_elems:
            raise ValueError("No <diagram> elements found in <mxfile>.")
        # Try each diagram page until one decodes.
        last_err: Optional[str] = None
        for diagram in diagram_elems:
            # Some exports embed <mxGraphModel> as a child element (not as encoded text)
            mx_child = diagram.find(".//mxGraphModel")
            if mx_child is not None:
                return ET.tostring(mx_child, encoding="unicode")

            # Otherwise, the diagram content is typically encoded in the diagram text
            decoded = _try_decode_diagram_payload(diagram.text or "")
            if decoded:
                return decoded

            # Sometimes text may be split; try itertext
            decoded2 = _try_decode_diagram_payload("".join(diagram.itertext()))
            if decoded2:
                return decoded2

            last_err = "Could not decode one <diagram> payload into <mxGraphModel>."

        raise ValueError(last_err or "Could not decode <diagram> payload into <mxGraphModel>.")

    # Case 3: Some exports have <diagram> as root
    if root.tag == "diagram":
        # Some exports embed <mxGraphModel> as a child element (not as encoded text)
        mx_child = root.find(".//mxGraphModel")
        if mx_child is not None:
            return ET.tostring(mx_child, encoding="unicode")

        decoded = _try_decode_diagram_payload(root.text or "")
        if decoded:
            return decoded

        decoded2 = _try_decode_diagram_payload("".join(root.itertext()))
        if decoded2:
            return decoded2

        raise ValueError("Could not decode <diagram> payload into <mxGraphModel>.")

    # Fallback: look for mxGraphModel anywhere
    mx = root.find(".//mxGraphModel")
    if mx is not None:
        return ET.tostring(mx, encoding="unicode")

    raise ValueError("Unable to locate <mxGraphModel> in input.")


# ----------------------------- mxGraph parsing -----------------------------

def parse_mxgraphmodel(mxgraph_xml_text: str):
    import xml.etree.ElementTree as ET
    return ET.fromstring(mxgraph_xml_text)


def _iter_cells(mxgraph_root) -> Iterable:
    for cell in mxgraph_root.findall(".//mxCell"):
        yield cell


def _cell_attr(cell, name: str) -> str:
    return cell.attrib.get(name, "")


def _cell_value(cell) -> str:
    """Extract label text from an mxCell value attribute (light HTML stripping)."""
    v = cell.attrib.get("value", "")
    v = re.sub(r"<br\s*/?>", "\n", v, flags=re.IGNORECASE)
    v = re.sub(r"</?[^>]+>", "", v)
    v = v.replace("&nbsp;", " ").strip()
    return v


def _cell_geometry(cell) -> Tuple[float, float, float, float]:
    geo = cell.find("./mxGeometry")
    if geo is None:
        return (0.0, 0.0, 0.0, 0.0)

    def _f(k: str) -> float:
        try:
            return float(geo.attrib.get(k, 0) or 0)
        except Exception:
            return 0.0

    return (_f("x"), _f("y"), _f("width"), _f("height"))


def _is_vertex(cell) -> bool:
    return _cell_attr(cell, "vertex") == "1"


# --------------------------- Source table extraction ------------------------

def _cluster_by_y(
    cells_with_geom: List[Tuple[float, float, float, float, str]],
    tolerance: float = 6.0,
):
    """
    Group cell labels into rows based on y coordinate.
    Returns list of rows, each row is list of (x, label).
    """
    sorted_cells = sorted(cells_with_geom, key=lambda t: (t[1], t[0]))
    rows: List[List[Tuple[float, str]]] = []
    row_ys: List[float] = []

    for x, y, w, h, label in sorted_cells:
        placed = False
        for i, ry in enumerate(row_ys):
            if abs(y - ry) <= tolerance:
                rows[i].append((x, label))
                placed = True
                break
        if not placed:
            row_ys.append(y)
            rows.append([(x, label)])

    for r in rows:
        r.sort(key=lambda t: t[0])

    return rows


def _looks_like_source_table_header(row_labels: List[str]) -> bool:
    norm = [re.sub(r"\s+", " ", x.strip().lower()) for x in row_labels if x.strip()]
    return ("column" in norm) and ("datatype" in norm)


def extract_source_tables_from_drawio(drawio_xml_text: str) -> List[SourceTable]:
    """
    Extract source tables from a draw.io diagram.

    Strategy:
      - Decode <mxGraphModel>
      - Find vertex cells with style containing 'shape=table'
      - Collect descendant vertices (table cells are often grandchildren under tableRow nodes)
      - Reconstruct a grid using geometry (y groups rows, x orders columns)
      - Identify header row containing 'Column' and 'Datatype'
      - Extract Column/Datatype/Size/Scale from subsequent rows
    """
    mx_xml = extract_mxgraphmodel_xml(drawio_xml_text)
    root = parse_mxgraphmodel(mx_xml)

    cells = list(_iter_cells(root))
    by_id: Dict[str, any] = {c.attrib.get("id", ""): c for c in cells if c.attrib.get("id")}

    # parent -> children (cells store "parent" attribute)
    children_by_parent: Dict[str, List[any]] = {}
    for c in cells:
        pid = c.attrib.get("parent", "")
        if pid:
            children_by_parent.setdefault(pid, []).append(c)

    def iter_descendant_vertices(parent_id: str) -> Iterable:
        """Yield all descendant mxCells (vertex=1) under a parent id (BFS)."""
        queue = [parent_id]
        seen = set()
        while queue:
            pid = queue.pop(0)
            for ch in children_by_parent.get(pid, []):
                cid = ch.attrib.get("id", "")
                if cid and cid not in seen:
                    seen.add(cid)
                    queue.append(cid)
                if _is_vertex(ch):
                    yield ch

    # Candidate: draw.io "table" stencil. IMPORTANT: exclude tableRow.
    # Some styles contain `shape=tableRow` which would match a naive substring search.
    candidates = []
    for c in cells:
        if not _is_vertex(c):
            continue
        style = c.attrib.get("style", "") or ""
        if "shape=tableRow" in style:
            continue
        # Require the exact shape key/value (with delimiters) to avoid false positives.
        if re.search(r"(^|;)shape=table($|;)", style):
            candidates.append(c)

    source_tables: List[SourceTable] = []

    for table_cell in candidates:
        tid = table_cell.attrib.get("id", "")
        if not tid:
            continue

        # Collect descendant vertices that actually hold cell labels (value) and geometry
        descendant_vertices = list(iter_descendant_vertices(tid))
        if not descendant_vertices:
            continue

        def abs_xy(cell_id: str) -> Tuple[float, float]:
            """Compute absolute x,y by summing mxGeometry x,y up the parent chain."""
            ax = 0.0
            ay = 0.0
            cur_id = cell_id
            seen = set()
            while cur_id and cur_id not in seen:
                seen.add(cur_id)
                cell = by_id.get(cur_id)
                if cell is None:
                    break
                x, y, _, _ = _cell_geometry(cell)
                ax += x
                ay += y
                cur_id = cell.attrib.get("parent", "")
                # Stop at top-level (parent '1' or '0' typically has no geometry impact)
                if cur_id in ("0", "1", ""):
                    break
            return ax, ay

        labeled_cells: List[Tuple[float, float, float, float, str]] = []
        for cc in descendant_vertices:
            style = cc.attrib.get("style", "") or ""
            # Ignore row container cells; we want actual table cells.
            if "shape=tableRow" in style:
                continue

            label = _cell_value(cc)
            x, y, w, h = _cell_geometry(cc)

            # Keep empty-label cells if they look like real table cells.
            # This prevents blank Scale cells from causing the next value (e.g., Target) to shift left.
            if not label.strip():
                if w <= 0 and h <= 0:
                    continue
                label = ""

            ax, ay = abs_xy(cc.attrib.get("id", ""))
            labeled_cells.append((ax, ay, w, h, label))

        if not labeled_cells:
            continue

        rows = _cluster_by_y(labeled_cells)

        header_idx = None
        header_labels = None
        for i, row in enumerate(rows):
            labels = [lbl for _, lbl in row if str(lbl).strip()]
            if _looks_like_source_table_header(labels):
                header_idx = i
                header_labels = labels
                break
        if header_idx is None or header_labels is None:
            continue

        header_norm = [re.sub(r"\s+", " ", h.strip().lower()) for h in header_labels]

        def idx_of(name: str) -> Optional[int]:
            try:
                return header_norm.index(name)
            except ValueError:
                return None

        col_i = idx_of("column")
        dt_i = idx_of("datatype")
        size_i = idx_of("size")
        scale_i = idx_of("scale")
        target_i = idx_of("target")

        if col_i is None or dt_i is None:
            continue

        table_name = _cell_value(table_cell) or tid or "source_table"

        columns: List[SourceColumn] = []
        for r in rows[header_idx + 1 :]:
            vals = [lbl for _, lbl in r]
            max_len = max(
                col_i if col_i is not None else 0,
                dt_i if dt_i is not None else 0,
                size_i if size_i is not None else 0,
                scale_i if scale_i is not None else 0,
                target_i if target_i is not None else 0,
            ) + 1
            if len(vals) < max_len:
                vals += [""] * (max_len - len(vals))

            col = (vals[col_i] if col_i < len(vals) else "").strip()
            dt = (vals[dt_i] if dt_i < len(vals) else "").strip()
            sz = (vals[size_i] if size_i is not None and size_i < len(vals) else "").strip()
            sc = (vals[scale_i] if scale_i is not None and scale_i < len(vals) else "").strip()
            tgt = (vals[target_i] if target_i is not None and target_i < len(vals) else "").strip()

            # Skip empty rows
            if not col and not dt and not sz and not sc and not tgt:
                continue
            if not col:
                continue

            columns.append(SourceColumn(column=col, datatype=dt, size=sz, scale=sc, target=tgt))

        if columns:
            source_tables.append(SourceTable(table_name=table_name, columns=columns))

    # As per current modelling convention: one source table per model file.
    if len(source_tables) > 1:
        raise ValueError(
            f"Expected exactly one source table in the model, but found {len(source_tables)}. "
            "If you intentionally model multiple source tables per file later, we can relax this rule."
        )
    return source_tables


# ----------------------------- XLSX generation -----------------------------

SOURCE_HEADERS = ["Table Schema", "Table Name", "Column", "Datatype", "Size", "Scale"]


TARGET_HEADERS = [
    "Table Type",
    "Subtype",
    "Table Name",
    "Column",
    "Datatype",
    "Size",
    "Scale",
    "Column Types",
    "Parent Table",
    "Relationship",
    "Relationship Name",
]

MAPPING_HEADERS = [
    "Source Table",
    "Source Column",
    "Target Table",
    "Target Column",
    "Mapping Set Name",
]


def write_source_xlsx(schema: str, source_tables: List[SourceTable], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Source"

    ws.append(SOURCE_HEADERS)

    for st in source_tables:
        for col in st.columns:
            ws.append([schema, st.table_name, col.column, col.datatype, col.size, col.scale])

    ws.freeze_panes = "A2"
    widths = {"A": 14, "B": 32, "C": 26, "D": 14, "E": 10, "F": 10}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------- Target helpers -----------------------------


def _style_kv(style: str) -> dict:
    d = {}
    for part in (style or "").split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            d[k] = v
    return d


def _clean_html_value(v: str) -> str:
    v = v or ""
    v = re.sub(r"<br\s*/?>", "\n", v, flags=re.IGNORECASE)
    v = re.sub(r"</div\s*>", "\n", v, flags=re.IGNORECASE)
    v = re.sub(r"<div[^>]*>", "", v, flags=re.IGNORECASE)
    v = re.sub(r"</p\s*>", "\n", v, flags=re.IGNORECASE)
    v = re.sub(r"<p[^>]*>", "", v, flags=re.IGNORECASE)
    v = re.sub(r"</?[^>]+>", "", v)
    v = v.replace("&nbsp;", " ")
    v = re.sub(r"\n+", "\n", v)
    return v.strip()


def _get_title_from_cell(cell) -> str:
    txt = _clean_html_value(cell.attrib.get("value", ""))
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
    return lines[0] if lines else ""


def _build_graph_indexes(mxgraph_root):
    cells = list(_iter_cells(mxgraph_root))
    by_id = {c.attrib.get("id", ""): c for c in cells if c.attrib.get("id")}
    children_by_parent: Dict[str, List[any]] = {}
    for c in cells:
        pid = c.attrib.get("parent", "")
        if pid:
            children_by_parent.setdefault(pid, []).append(c)
    return cells, by_id, children_by_parent


def _derive_relationship_name(hub_name: str) -> str:
    return hub_name[2:] if hub_name.lower().startswith("h_") else hub_name


def _hub_column_rows(hub_name: str) -> List[Tuple[str, str, str, str, str]]:
    concept = _derive_relationship_name(hub_name)
    return [
        ("bkcc", "varchar", "100", "", "BKCC"),
        (f"bk_{concept}", "varchar", "100", "", "Business key"),
    ]


def _link_column_rows(related_hubs: List[str]) -> List[Tuple[str, str, str, str, str]]:
    rows = []
    for hub in related_hubs:
        concept = _derive_relationship_name(hub)
        rows.append((f"bkcc_{concept}", "varchar", "100", "", "Link BKCC"))
        rows.append((f"bk_{concept}", "varchar", "100", "", "Link business key"))
    return rows


def write_target_xlsx(target_rows: List[List[str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Target"

    ws.append(TARGET_HEADERS)
    for r in target_rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    widths = {
        "A": 14,
        "B": 18,
        "C": 34,
        "D": 26,
        "E": 14,
        "F": 10,
        "G": 10,
        "H": 20,
        "I": 34,
        "J": 24,
        "K": 24,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_mapping_xlsx(mapping_rows: List[List[str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    # Match the example workbook which uses the default sheet name.
    ws.title = "Sheet1"

    ws.append(MAPPING_HEADERS)
    for r in mapping_rows:
        ws.append(r)

    ws.freeze_panes = "A2"
    widths = {"A": 34, "B": 26, "C": 34, "D": 26, "E": 26}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------- Public function -----------------------------

def generate_source_metadata(model_filename: str, schema: str) -> Path:
    """
    Generate a brand-new IRiS Source metadata XLSX for a given draw.io XML model file.

    Looks for input in:  ./model/<model_filename>
    Writes output to:    ./output/<MODEL_NAME>/source_<MODEL_NAME>.xlsx

    Returns the output file path.
    """
    script_dir = Path(__file__).resolve().parent
    model_dir = script_dir / "model"
    out_root = script_dir / "output"

    model_path = model_dir / model_filename
    if not model_path.exists():
        raise FileNotFoundError(f"Model file not found: {model_path}")

    # Determine model name from filename stem (strip common multi-extensions)
    name = model_path.name
    model_name = re.sub(r"\.drawio\.xml$", "", name, flags=re.IGNORECASE)
    model_name = re.sub(r"\.xml$", "", model_name, flags=re.IGNORECASE)
    model_name = Path(model_name).stem

    out_dir = out_root / model_name
    out_file = out_dir / f"source_{model_name}.xlsx"

    drawio_text = _read_text(model_path)
    source_tables = extract_source_tables_from_drawio(drawio_text)
    if not source_tables:
        raise ValueError(
            "No source tables found. Expected at least one draw.io table shape with header cells "            "including 'Column' and 'Datatype'."
        )

    write_source_xlsx(schema=schema, source_tables=source_tables, out_path=out_file)
    return out_file


# ----------------------------- Target generator -----------------------------


def generate_target_metadata(model_filename: str) -> Path:
    """Generate IRiS Target metadata XLSX from a draw.io XML model file."""
    script_dir = Path(__file__).resolve().parent
    model_dir = script_dir / "model"
    out_root = script_dir / "output"

    model_path = model_dir / model_filename
    if not model_path.exists():
        raise FileNotFoundError(f"Model file not found: {model_path}")

    # Determine model name from filename stem (strip common multi-extensions)
    name = model_path.name
    model_name = re.sub(r"\.drawio\.xml$", "", name, flags=re.IGNORECASE)
    model_name = re.sub(r"\.xml$", "", model_name, flags=re.IGNORECASE)
    model_name = Path(model_name).stem

    out_dir = out_root / model_name
    out_file = out_dir / f"target_{model_name}.xlsx"

    drawio_text = _read_text(model_path)
    mx_xml = extract_mxgraphmodel_xml(drawio_text)

    import xml.etree.ElementTree as ET

    mx_root = ET.fromstring(mx_xml)

    # draw.io can store template shapes as `<object id=... label=...>` wrapping an inner `<mxCell>`.
    # Edges will reference the object id, but the inner mxCell may have no id. Inject synthetic mxCell
    # nodes with the object id so the rest of the parsing works uniformly.
    root_node = mx_root.find(".//root")
    if root_node is not None:
        for obj in list(mx_root.findall(".//object")):
            obj_id = obj.attrib.get("id", "")
            obj_label = obj.attrib.get("label", "")
            inner = obj.find("./mxCell")
            if not obj_id or inner is None:
                continue

            # Create a synthetic mxCell with the object id and label.
            synth = ET.Element("mxCell")
            synth.attrib.update(inner.attrib)
            synth.attrib["id"] = obj_id
            synth.attrib["value"] = obj_label
            # Ensure it's treated as a vertex unless explicitly set otherwise.
            synth.attrib.setdefault("vertex", "1")

            # Copy geometry if present.
            geo = inner.find("./mxGeometry")
            if geo is not None:
                synth.append(ET.fromstring(ET.tostring(geo, encoding="utf-8")))

            root_node.append(synth)

            # Also, if the object carries a Source Extract Date attribute, store it on the synthetic cell
            # so our existing scan logic can pick it up.
            for k, v in obj.attrib.items():
                if k.lower().replace("_", "").replace(" ", "") == "sourceextractdate":
                    # Store as an extra attribute for downstream scanning.
                    synth.attrib["_source_extract_date"] = v

    cells, by_id, children_by_parent = _build_graph_indexes(mx_root)

    # Detect Dependent Child Satellite and Multi-active Satellite template fillColor (optional but preferred).
    dependent_child_fill: Optional[str] = None
    multiactive_fill: Optional[str] = None
    template_path = script_dir / "IRiS_DV_Modelling.xml"
    if template_path.exists():
        tpl_txt = template_path.read_text(encoding="utf-8", errors="replace")
        m = re.search(r"<mxlibrary>(.*)</mxlibrary>", tpl_txt, flags=re.DOTALL)
        if m:
            try:
                arr = json.loads(m.group(1))
            except Exception:
                arr = []
            for item in arr:
                title = (item.get("title", "") or "").strip().lower()
                want_dep = title == "dependent child satellite"
                want_ma = title in ("multiactive satellite", "multi-active satellite", "multi active satellite")

                if not (want_dep or want_ma):
                    continue

                xml_s = html.unescape(item.get("xml", "") or "")
                try:
                    tr = ET.fromstring(xml_s)
                except Exception:
                    continue
                verts = [c for c in tr.findall(".//mxCell") if c.attrib.get("vertex") == "1"]
                if not verts:
                    continue

                def area(c):
                    g = c.find("./mxGeometry")
                    if g is None:
                        return 0.0
                    return float(g.attrib.get("width", "0") or 0) * float(g.attrib.get("height", "0") or 0)

                main = sorted(verts, key=area, reverse=True)[0]
                fill = _style_kv(main.attrib.get("style", "")).get("fillColor")
                if not fill:
                    continue

                if want_dep and dependent_child_fill is None:
                    dependent_child_fill = fill
                if want_ma and multiactive_fill is None:
                    multiactive_fill = fill

                # Stop once we've found both (if both exist in the library)
                if dependent_child_fill is not None and multiactive_fill is not None:
                    break
    # Extract the (single) Source table and its columns (incl. in-memory `target` values).
    source_tables = extract_source_tables_from_drawio(drawio_text)
    source_table = source_tables[0] if source_tables else None

    # Build a quick index of Source pseudo-mapping rows by Target DV object.
    source_by_target: Dict[str, List[SourceColumn]] = {}
    if source_table is not None:
        for sc in source_table.columns:
            tgt = (sc.target or "").strip()
            if not tgt:
                continue
            source_by_target.setdefault(tgt.lower(), []).append(sc)

    def _pick_source_row_for_hub(hub_name: str, want_bkcc: bool) -> Optional[SourceColumn]:
        """Pick a representative source column row for a hub: BKCC row or Business Key row."""
        rows = source_by_target.get((hub_name or "").strip().lower(), [])
        if not rows:
            return None
        if want_bkcc:
            # Prefer a row whose source column name starts with bkcc
            for r in rows:
                if (r.column or "").strip().lower().startswith("bkcc"):
                    return r
            return None
        # Business key: first non-bkcc row (e.g. customerid/serviceid/etc)
        for r in rows:
            if not (r.column or "").strip().lower().startswith("bkcc"):
                return r
        return None

    def _lookup_source_type_for_target(table_type: str, table_name: str, target_column: str) -> Tuple[str, str, str]:
        """Return (datatype, size, scale) for a DV Target column using the Source pseudo-mapping."""
        ttype = (table_type or "").strip().lower()
        tname = (table_name or "").strip()
        tcol = (target_column or "").strip()

        if ttype == "hub":
            # Hub target columns are modelled as bkcc + bk_<concept>, but source columns may be bkcc_<concept> + <raw bk>.
            if tcol.lower() == "bkcc":
                r = _pick_source_row_for_hub(tname, want_bkcc=True)
            else:
                r = _pick_source_row_for_hub(tname, want_bkcc=False)
            if r is None:
                return ("", "", "")
            return (r.datatype or "", r.size or "", r.scale or "")

        if ttype == "link":
            # Link target columns are modelled as bkcc_<concept> and bk_<concept>.
            # Derive types from the corresponding hub’s source mapping.
            m = re.match(r"^(bkcc_|bk_)(.+)$", tcol, flags=re.IGNORECASE)
            if not m:
                return ("", "", "")
            prefix = m.group(1).lower()
            concept = (m.group(2) or "").strip().lower()
            if not concept:
                return ("", "", "")
            hub_name = f"h_{concept}"
            want_bkcc = prefix.startswith("bkcc_")
            r = _pick_source_row_for_hub(hub_name, want_bkcc=want_bkcc)
            if r is None:
                return ("", "", "")
            return (r.datatype or "", r.size or "", r.scale or "")

        # Satellites already come directly from Source rows filtered by Target == satellite name.
        return ("", "", "")

    # Satellite subtype detection
    satellite_subtypes: Dict[str, str] = {}  # satellite_name -> subtype

    def _nearest_fill_color(cell_id: str) -> Optional[str]:
        """Walk upwards from a cell id and return the first non-empty fillColor."""
        for aid in _walk_ancestors(cell_id):
            c = by_id.get(aid)
            if c is None:
                continue
            fill = _style_kv(c.attrib.get("style", "")).get("fillColor")
            if fill:
                return fill
        return None

    # Identify DV objects by name from any vertex label.
    # We map both the label cell id and its ancestors to the same DV object name/type.
    obj_by_id: Dict[str, Tuple[str, str]] = {}  # cell_id -> (name, type)
    objects: Dict[str, str] = {}  # name -> type

    def _walk_ancestors(cell_id: str) -> Iterable[str]:
        cur = cell_id
        seen = set()
        while cur and cur not in seen:
            if cur in ("0", "1"):
                break
            seen.add(cur)
            yield cur
            cell = by_id.get(cur)
            if cell is None:
                break
            cur = cell.attrib.get("parent", "")

    for c in cells:
        if c.attrib.get("vertex") != "1":
            continue
        if c.attrib.get("edge") == "1":
            continue
        title = _get_title_from_cell(c)
        if not title:
            continue
        low = title.lower()
        if low.startswith("h_"):
            t = "Hub"
        elif low.startswith("l_"):
            t = "Link"
        elif low.startswith("s_"):
            t = "Satellite"
        else:
            continue

        objects[title] = t
        cid = c.attrib.get("id", "")
        if cid:
            # Subtype detection: Dependent Child Satellite / Multi-active Satellite
            if t == "Satellite":
                low_name = title.lower()
                subtype = ""

                # Dependent child satellite
                if low_name.startswith("s_dc_"):
                    subtype = "Dependent Child Satellite"
                else:
                    fill = _nearest_fill_color(cid)
                    if dependent_child_fill and fill and fill.lower() == dependent_child_fill.lower():
                        subtype = "Dependent Child Satellite"

                # Multi-active satellite (only if not already marked as dependent child)
                if not subtype:
                    if low_name.startswith("s_ma_"):
                        subtype = "multi active"
                    else:
                        fill = _nearest_fill_color(cid)
                        if multiactive_fill and fill and fill.lower() == multiactive_fill.lower():
                            subtype = "multi active"

                if subtype:
                    satellite_subtypes[title] = subtype

            for aid in _walk_ancestors(cid):
                a_cell = by_id.get(aid)
                if a_cell is None:
                    continue
                # Only map vertex cells to avoid incorrectly assigning layer/page containers.
                if a_cell.attrib.get("vertex") != "1":
                    continue
                # Don't overwrite if already mapped (first wins)
                obj_by_id.setdefault(aid, (title, t))

    # Precompute Source Extract Date annotations by Satellite.
    # The annotation cell can be a sibling of the name label, so resolve ownership via `obj_by_id` + ancestor walk.
    sed_cols_by_sat: Dict[str, set] = {}

    # Precompute Dependent Child Attribute annotations by Satellite.
    # Expected label example in the diagram: 'Dependent_Child_Attribute: <column_name>'
    depchild_cols_by_sat: Dict[str, set] = {}

    def _resolve_obj_for_cell(cell_id: str) -> Optional[Tuple[str, str]]:
        if not cell_id:
            return None
        # Prefer exact match first
        if cell_id in obj_by_id:
            return obj_by_id[cell_id]
        # Then prefer the nearest ancestor that is mapped
        for aid in _walk_ancestors(cell_id):
            obj = obj_by_id.get(aid)
            if obj:
                return obj
        return None

    for c in cells:
        cid = c.attrib.get("id", "")
        if not cid:
            continue

        raw = c.attrib.get("value", "") or ""

        # Some templates store Source Extract Date as an object/cell attribute.
        injected_sed = (c.attrib.get("_source_extract_date", "") or "").strip()

        # Also look for any attribute key that normalises to 'sourceextractdate'
        for k, v in c.attrib.items():
            if k.lower().replace("_", "").replace(" ", "") == "sourceextractdate":
                if (v or "").strip():
                    injected_sed = (v or "").strip()

        # If we have an injected/attribute value, record it even if there is no visible text value.
        if injected_sed:
            obj = _resolve_obj_for_cell(cid)
            if obj:
                obj_name, obj_type = obj
                if obj_type == "Satellite":
                    sed_cols_by_sat.setdefault(obj_name, set()).add(injected_sed.strip().lower())

        txt = _clean_html_value(raw)
        if not txt:
            continue

        for line in [ln.strip() for ln in txt.splitlines() if ln.strip()]:
            norm = re.sub(r"[_\s]+", " ", line).strip().lower()

            # -------- Source Extract Date --------
            if norm.startswith("source extract date"):
                m = re.match(r"^source[_\s]*extract[_\s]*date\s*[:=]\s*(.+)$", line, flags=re.IGNORECASE)
                col = m.group(1).strip() if m else ""
                if not col:
                    col = re.sub(r"^source[_\s]*extract[_\s]*date\s*", "", line, flags=re.IGNORECASE).strip()
                col = re.sub(r"\s+", " ", col).strip()
                if col:
                    obj = _resolve_obj_for_cell(cid)
                    if obj and obj[1] == "Satellite":
                        sed_cols_by_sat.setdefault(obj[0], set()).add(col.strip().lower())
                continue

            # -------- Dependent Child Attribute --------
            if norm.startswith("dependent child attribute"):
                m = re.match(r"^dependent[_\s]*child[_\s]*attribute\s*[:=]\s*(.+)$", line, flags=re.IGNORECASE)
                col = m.group(1).strip() if m else ""
                if not col:
                    col = re.sub(r"^dependent[_\s]*child[_\s]*attribute\s*", "", line, flags=re.IGNORECASE).strip()
                col = re.sub(r"\s+", " ", col).strip()
                if col:
                    obj = _resolve_obj_for_cell(cid)
                    if obj and obj[1] == "Satellite":
                        depchild_cols_by_sat.setdefault(obj[0], set()).add(col.strip().lower())
                continue

    # Resolve edges to Link-Hub and Satellite-Parent using ancestor-walk mapping.
    link_to_hubs: Dict[str, List[Tuple[str, str]]] = {}  # link_name -> [(hub_name, label)]
    sat_to_parent: Dict[str, str] = {}  # sat_name -> parent_name

    for e in cells:
        if e.attrib.get("edge") != "1":
            continue
        src = e.attrib.get("source", "")
        tgt = e.attrib.get("target", "")
        if not src or not tgt:
            continue

        src_obj = obj_by_id.get(src)
        tgt_obj = obj_by_id.get(tgt)

        # If the edge connects to an inner cell, try walking ancestors.
        if src_obj is None:
            for aid in _walk_ancestors(src):
                src_obj = obj_by_id.get(aid)
                if src_obj:
                    break
        if tgt_obj is None:
            for aid in _walk_ancestors(tgt):
                tgt_obj = obj_by_id.get(aid)
                if tgt_obj:
                    break

        if not src_obj or not tgt_obj:
            continue

        src_name, src_type = src_obj
        tgt_name, tgt_type = tgt_obj
        edge_label = _clean_html_value(e.attrib.get("value", "")) if e.attrib.get("value") else ""

        if src_type == "Link" and tgt_type == "Hub":
            link_to_hubs.setdefault(src_name, []).append((tgt_name, edge_label))
        elif tgt_type == "Link" and src_type == "Hub":
            link_to_hubs.setdefault(tgt_name, []).append((src_name, edge_label))

        if src_type == "Satellite" and tgt_type in ("Hub", "Link"):
            sat_to_parent[src_name] = tgt_name
        elif tgt_type == "Satellite" and src_type in ("Hub", "Link"):
            sat_to_parent[tgt_name] = src_name

    # Build Target rows.
    target_rows: List[List[str]] = []

    # Hubs
    for nm in sorted([n for n, t in objects.items() if t == "Hub"]):
        for col, dt, sz, sc, ctype in _hub_column_rows(nm):
            # Per convention: all BKCC + Business key target columns are varchar(100)
            dt, sz, sc = "varchar", "100", ""
            target_rows.append(["Hub", "", nm, col, dt, sz, sc, ctype, "", "", ""])

    # Links
    for nm in sorted([n for n, t in objects.items() if t == "Link"]):
        rels = link_to_hubs.get(nm, [])
        # Relationship names: prefer edge labels; else hub concept name
        counts: Dict[str, int] = {}
        rel_entries: List[Tuple[str, str]] = []  # (hub_name, relationship_name)
        for hub_name, edge_label in sorted(rels, key=lambda x: x[0]):
            base = edge_label.strip() if edge_label.strip() else _derive_relationship_name(hub_name)
            counts[base] = counts.get(base, 0) + 1
            rel_name = base if counts[base] == 1 else f"{base}_{counts[base]}"
            rel_entries.append((hub_name, rel_name))

        related_hubs = [h for h, _ in rel_entries]
        for col, dt, sz, sc, ctype in _link_column_rows(related_hubs):
            rel = ""
            rel_name = ""
            m2 = re.match(r"^(?:bkcc_|bk_)(.+)$", col, flags=re.IGNORECASE)
            if m2:
                concept = m2.group(1)
                for hub_name, rname in rel_entries:
                    if _derive_relationship_name(hub_name).lower() == concept.lower():
                        rel = hub_name
                        rel_name = rname
                        break
            # Per convention: all Link BKCC + Link business key target columns are varchar(100)
            dt, sz, sc = "varchar", "100", ""
            target_rows.append(["Link", "", nm, col, dt, sz, sc, ctype, "", rel, rel_name])

    def _iter_descendants(start_id: str) -> Iterable[str]:
        """Yield descendant cell ids (including the start id)."""
        queue = [start_id]
        seen = set()
        while queue:
            cid = queue.pop(0)
            if not cid or cid in seen:
                continue
            seen.add(cid)
            yield cid
            for ch in children_by_parent.get(cid, []):
                queue.append(ch.attrib.get("id", ""))

    def _extract_source_extract_date_columns_for_satellite(sat_name: str) -> List[str]:
        """Return list of column names referenced by a 'Source Extract Date' annotation inside the satellite shape."""
        sat_name = (sat_name or "").strip()
        if not sat_name:
            return []

        # Find any vertex cell whose title equals the satellite name.
        start_ids = []
        for c in cells:
            if c.attrib.get("vertex") != "1" or c.attrib.get("edge") == "1":
                continue
            if _get_title_from_cell(c) == sat_name:
                cid = c.attrib.get("id", "")
                if cid:
                    start_ids.append(cid)

        if not start_ids:
            return []

        cols: List[str] = []
        for sid in start_ids:
            for did in _iter_descendants(sid):
                cell = by_id.get(did)
                if cell is None:
                    continue
                raw = cell.attrib.get("value", "")
                if not raw:
                    continue
                txt = _clean_html_value(raw)
                if not txt:
                    continue

                # Accept variants like:
                #   Source_Extract_Date: servicedate
                #   Source Extract Date = servicedate
                #   source extract date servicedate
                for line in [ln.strip() for ln in txt.splitlines() if ln.strip()]:
                    norm = re.sub(r"[_\s]+", " ", line).strip().lower()
                    if not norm.startswith("source extract date"):
                        continue

                    # Try to parse a referenced column name after ':' or '='
                    m = re.match(r"^source[_\s]*extract[_\s]*date\s*[:=]\s*(.+)$", line, flags=re.IGNORECASE)
                    if m:
                        col = m.group(1).strip()
                    else:
                        # Fallback: take trailing token(s) after the phrase
                        rest = re.sub(r"^source[_\s]*extract[_\s]*date\s*", "", line, flags=re.IGNORECASE).strip()
                        col = rest

                    # Clean common noise
                    col = re.sub(r"\s+", " ", col).strip()
                    # If the value looks like a '<something>: <col>' we already handled, else allow a single identifier.
                    if col:
                        cols.append(col)

        # Normalise: return unique list preserving order
        seen = set()
        out = []
        for c in cols:
            key = c.strip().lower()
            if not key or key in seen:
                continue
            seen.add(key)
            out.append(c.strip())
        return out

    # Satellites
    for nm in sorted([n for n, t in objects.items() if t == "Satellite"]):
        parent = sat_to_parent.get(nm, "")
        subtype = satellite_subtypes.get(nm, "")

        # Columns referenced by the 'Source Extract Date' annotation inside the satellite shape.
        sed_cols_l = sed_cols_by_sat.get(nm, set())
        depchild_cols_l = depchild_cols_by_sat.get(nm, set())

        # Your rule: satellite attributes are all source attributes WHERE source.Target == satellite name.
        sat_cols: List[SourceColumn] = []
        if source_table is not None:
            sat_cols = [c for c in source_table.columns if (c.target or "").strip() == nm]

        for c in sat_cols:
            col_l = (c.column or "").strip().lower()
            if col_l in depchild_cols_l:
                col_type = "Dependent child key"
                dt = "varchar"
                sz = "100"
                sc = ""
            elif col_l in sed_cols_l:
                col_type = "Source extract date"
                dt = c.datatype
                sz = c.size
                sc = c.scale
            else:
                col_type = "Changing attribute"
                dt = c.datatype
                sz = c.size
                sc = c.scale

            target_rows.append(["Satellite", subtype, nm, c.column, dt, sz, sc, col_type, parent, "", ""])

        # If none matched, still include a placeholder row so the sat exists.
        if not sat_cols:
            target_rows.append(["Satellite", subtype, nm, "", "", "", "", "Changing attribute", parent, "", ""])

    write_target_xlsx(target_rows, out_file)
    return out_file


def generate_mapping_metadata(model_filename: str) -> Path:
    """Generate IRiS Mapping metadata XLSX from a draw.io XML model file."""
    script_dir = Path(__file__).resolve().parent
    model_dir = script_dir / "model"
    out_root = script_dir / "output"

    model_path = model_dir / model_filename
    if not model_path.exists():
        raise FileNotFoundError(f"Model file not found: {model_path}")

    # Determine model name from filename stem (strip common multi-extensions)
    name = model_path.name
    model_name = re.sub(r"\.drawio\.xml$", "", name, flags=re.IGNORECASE)
    model_name = re.sub(r"\.xml$", "", model_name, flags=re.IGNORECASE)
    model_name = Path(model_name).stem

    out_dir = out_root / model_name
    out_file = out_dir / f"mapping_{model_name}.xlsx"

    drawio_text = _read_text(model_path)

    # Parse model for Link-Hub relationships.
    mx_xml = extract_mxgraphmodel_xml(drawio_text)

    import xml.etree.ElementTree as ET

    mx_root = ET.fromstring(mx_xml)

    # Inject synthetic mxCells for `<object id=... label=...>` wrappers so edge endpoints resolve.
    root_node = mx_root.find(".//root")
    if root_node is not None:
        for obj in list(mx_root.findall(".//object")):
            obj_id = obj.attrib.get("id", "")
            obj_label = obj.attrib.get("label", "")
            inner = obj.find("./mxCell")
            if not obj_id or inner is None:
                continue

            synth = ET.Element("mxCell")
            synth.attrib.update(inner.attrib)
            synth.attrib["id"] = obj_id
            synth.attrib["value"] = obj_label
            synth.attrib.setdefault("vertex", "1")

            geo = inner.find("./mxGeometry")
            if geo is not None:
                synth.append(ET.fromstring(ET.tostring(geo, encoding="utf-8")))

            root_node.append(synth)

    cells, by_id, children_by_parent = _build_graph_indexes(mx_root)

    def _walk_ancestors(cell_id: str) -> Iterable[str]:
        cur = cell_id
        seen = set()
        while cur and cur not in seen:
            if cur in ("0", "1"):
                break
            seen.add(cur)
            yield cur
            cell = by_id.get(cur)
            if cell is None:
                break
            cur = cell.attrib.get("parent", "")

    # Map cell ids (and their vertex ancestors) to DV object name/type.
    obj_by_id: Dict[str, Tuple[str, str]] = {}
    objects: Dict[str, str] = {}

    for c in cells:
        if c.attrib.get("vertex") != "1" or c.attrib.get("edge") == "1":
            continue
        title = _get_title_from_cell(c)
        if not title:
            continue
        low = title.lower()
        if low.startswith("h_"):
            t = "Hub"
        elif low.startswith("l_"):
            t = "Link"
        elif low.startswith("s_"):
            t = "Satellite"
        else:
            continue

        objects[title] = t
        cid = c.attrib.get("id", "")
        if cid:
            for aid in _walk_ancestors(cid):
                a_cell = by_id.get(aid)
                if a_cell is None:
                    continue
                if a_cell.attrib.get("vertex") != "1":
                    continue
                obj_by_id.setdefault(aid, (title, t))

    # Build Link -> connected Hubs using edges.
    link_to_hubs: Dict[str, List[str]] = {}

    for e in cells:
        if e.attrib.get("edge") != "1":
            continue
        src = e.attrib.get("source", "")
        tgt = e.attrib.get("target", "")
        if not src or not tgt:
            continue

        src_obj = obj_by_id.get(src)
        tgt_obj = obj_by_id.get(tgt)

        if src_obj is None:
            for aid in _walk_ancestors(src):
                src_obj = obj_by_id.get(aid)
                if src_obj:
                    break
        if tgt_obj is None:
            for aid in _walk_ancestors(tgt):
                tgt_obj = obj_by_id.get(aid)
                if tgt_obj:
                    break

        if not src_obj or not tgt_obj:
            continue

        src_name, src_type = src_obj
        tgt_name, tgt_type = tgt_obj

        if src_type == "Link" and tgt_type == "Hub":
            link_to_hubs.setdefault(src_name, []).append(tgt_name)
        elif tgt_type == "Link" and src_type == "Hub":
            link_to_hubs.setdefault(tgt_name, []).append(src_name)

    # Normalise hub lists
    for k in list(link_to_hubs.keys()):
        hubs = sorted(set(link_to_hubs[k]))
        link_to_hubs[k] = hubs

    # Mapping comes from the grey Source table pseudo-mapping: each row has Source Column + Target DV object.
    source_tables = extract_source_tables_from_drawio(drawio_text)
    if not source_tables:
        raise ValueError("No source tables found; cannot generate mapping.")
    source_table = source_tables[0]

    source_by_target: Dict[str, List[SourceColumn]] = {}
    for sc in source_table.columns:
        tgt = (sc.target or "").strip()
        if not tgt:
            continue
        source_by_target.setdefault(tgt.lower(), []).append(sc)

    def _pick_source_col_for_hub(hub_name: str, want_bkcc: bool) -> Optional[str]:
        rows = source_by_target.get((hub_name or "").strip().lower(), [])
        if not rows:
            return None
        if want_bkcc:
            for r in rows:
                if (r.column or "").strip().lower().startswith("bkcc"):
                    return (r.column or "").strip()
            return None
        for r in rows:
            if not (r.column or "").strip().lower().startswith("bkcc"):
                return (r.column or "").strip()
        return None

    source_table_name = source_table.table_name
    mapping_set_name = model_name

    def _target_type(target_table: str) -> str:
        t = (target_table or "").strip().lower()
        if t.startswith("h_"):
            return "hub"
        if t.startswith("l_"):
            return "link"
        if t.startswith("s_"):
            return "sat"
        return ""

    def _hub_concept(hub_name: str) -> str:
        hn = (hub_name or "").strip()
        return hn[2:] if hn.lower().startswith("h_") else hn

    def _derive_target_column(src_col: str, tgt_table: str) -> str:
        """Derive the target column name given a source column and the target DV object."""
        src_col = (src_col or "").strip()
        tgt_table = (tgt_table or "").strip()
        ttype = _target_type(tgt_table)

        if ttype == "hub":
            # Hub has standardised column names: bkcc + bk_<concept>
            if src_col.lower().startswith("bkcc_") or src_col.lower() == "bkcc":
                return "bkcc"
            return f"bk_{_hub_concept(tgt_table)}"

        if ttype in ("link", "sat"):
            # Links and Satellites keep their column names.
            return src_col

        # Unknown target type; pass through.
        return src_col

    mapping_rows: List[List[str]] = []
    seen_rows: set = set()  # tuples of (src_table, src_col, tgt_table, tgt_col, set_name)

    def _add_row(src_table: str, src_col: str, tgt_table: str, tgt_col: str, set_name: str) -> None:
        key = (src_table, src_col, tgt_table, tgt_col, set_name)
        if key in seen_rows:
            return
        seen_rows.add(key)
        mapping_rows.append([src_table, src_col, tgt_table, tgt_col, set_name])

    for sc in source_table.columns:
        tgt = (sc.target or "").strip()
        if not tgt:
            continue

        src_col = (sc.column or "").strip()
        tgt_col = _derive_target_column(src_col, tgt)

        _add_row(source_table_name, src_col, tgt, tgt_col, mapping_set_name)

    # Additional Link mappings:
    # If the source pseudo-mapping only targets Hubs, still emit mappings into Links based on link-hub relationships.
    for link_name, hubs in sorted(link_to_hubs.items(), key=lambda x: x[0]):
        for hub_name in hubs:
            concept = _derive_relationship_name(hub_name)

            src_bkcc = _pick_source_col_for_hub(hub_name, want_bkcc=True)
            if src_bkcc:
                _add_row(source_table_name, src_bkcc, link_name, f"bkcc_{concept}", mapping_set_name)

            src_bk = _pick_source_col_for_hub(hub_name, want_bkcc=False)
            if src_bk:
                _add_row(source_table_name, src_bk, link_name, f"bk_{concept}", mapping_set_name)

    write_mapping_xlsx(mapping_rows, out_file)
    return out_file


# ----------------------------- CLI (optional) ------------------------------

def _list_model_files(model_dir: Path) -> List[Path]:
    if not model_dir.exists():
        return []
    return sorted([p for p in model_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xml"])


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate IRiS Source + Target metadata XLSX from draw.io XML models.")
    parser.add_argument("--schema", required=True, help="Schema to write into the Source XLSX (e.g., landing).")
    parser.add_argument("--model", help="Model filename under ./model (e.g., SERVICE_SERVICESSYSTEM.drawio.xml).")
    parser.add_argument("--all", action="store_true", help="Process all .xml files in ./model.")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging.")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    script_dir = Path(__file__).resolve().parent
    model_dir = script_dir / "model"

    try:
        if args.all:
            model_files = _list_model_files(model_dir)
            if not model_files:
                logging.error("No model files found under ./model.")
                return 2
            for p in model_files:
                logging.info("Processing model: %s", p.name)
                out = generate_source_metadata(model_filename=p.name, schema=args.schema)
                logging.info("Wrote: %s", out)
                out_t = generate_target_metadata(model_filename=p.name)
                logging.info("Wrote: %s", out_t)
                out_m = generate_mapping_metadata(model_filename=p.name)
                logging.info("Wrote: %s", out_m)
            return 0

        if not args.model:
            logging.error("Provide --model <filename> or use --all.")
            return 2

        out = generate_source_metadata(model_filename=args.model, schema=args.schema)
        logging.info("Wrote: %s", out)
        out_t = generate_target_metadata(model_filename=args.model)
        logging.info("Wrote: %s", out_t)
        out_m = generate_mapping_metadata(model_filename=args.model)
        logging.info("Wrote: %s", out_m)
        return 0

    except Exception as e:
        logging.error("%s", e)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
